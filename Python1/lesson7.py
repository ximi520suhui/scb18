'''
接口自动化测试
1、excel测试用例准备ok，代码可以自动去读取用例数据---read_case（filename,sheetname）读测试用例的函数
2、执行接口测试，得到响应结果                    ----api_fun(url，data)
3、断言：响应结果跟预期结果比对，判断接口是否通过   ----通过/不通过
4、最终执行的结果写到测试用例的文档---excel表格----write_result（(filename,sheetname,row,column,final_result）

'''
import requests
import openpyxl
def read_case(filename,sheetname):
    wb=openpyxl.load_workbook(filename)    # 加载工作簿，打开一个excel文件
    sheet=wb[sheetname]     # 打开某个表单
    rows=sheet.max_row
    case_list=[]        # 新建一个空列表，存放for循环依次读取到的测试数据
    for i in range(2, rows + 1):
        data_dict = dict(                      # 把上面返回的数据，打包成一个字典形式的包
        case_id=sheet.cell(row=i, column=1).value,
        url=sheet.cell(row=i, column=5).value,  # 获取url的值
        data=sheet.cell(row=i, column=6).value,  # 获取测试数据data的值
        expect=sheet.cell(row=i, column=7).value
        )
        # print(data_dict)     # 输出case_id：对应的参数,url：对应的参数,data：对应的参数,expect：对应的参数（字典形式更容易看懂）
        case_list.append(data_dict)
        # print(case_list)       # 输出的数据是在上一次循环的后面追加1次（case_id:1;case_id:1,case_2;case_id:1,case_2,case_id:3...）
    return case_list

def api_fun(url,data):
    headers={'X-Lemonban-Media-Type':'lemonban.v2','Content-Type':'application/json'}   # 请求头
    res=requests.post(url=url,json=data,headers=headers).json()
    return res

def write_result(filename,sheetname,row,column,final_result):
    wb=openpyxl.load_workbook(filename)    # 加载工作簿，打开一个excel文件
    sheet=wb[sheetname]                    # 打开某一个表单
    sheet.cell(row=row,column=column).value=final_result
    wb.save(filename)


#断言：实际结果--预期结果
# cases=read_case('test_case_api.xlsx','register')    #调用函数，设置变量接收返回值
# print(cases)     #打印返回值
# for case in cases:
#     url=case['url']         #获取url的实际参数
#     data=case['data']         #获取消息体的实际参数，输出数据为字符串格式
#     # print(data)
#     # print(type(data))
#     expect=case['expected']     #获取期望值的参数
#     real_result=api_fun(url=url,data=data)     #连接接口，传数据，得到结果
#     print(real_result)

# eval()  ---运行被字符串包裹着的表达式
# dict0=(eval('{"mobile_phone":"13652440101","pwd":"12345678","type":1,"reg_name":"34254sdfs"}'))
# print(dict0)
# print(type(dict0))
# print(eval('2+6'))         # 直接输出8


# cases=read_case('test_case_api.xlsx','register')
# for case in cases:           # 直接在上面获取到的表格生成的list数据，循环取出要的单个数据,,得出的数据都是字符串数据类型
#     case_id=case['case_id']
#     url=case['url']           # 定义变量名为url，获取表格里的’url‘的值
#     data=eval(case['data'])       # 获取data的值,,把字符串转成字典
#     expect=eval(case['expect'])    #'''获取expect的值,(他的值不是在表格直接取，而是在上面def read_case下的循环里定义的expect值，也是字符串形式，需要转换成字典)'''
#     # print(url,data,expect)
#     # print(case_id, expect)
#     real_result=api_fun(url=url,data=data)
#     print(real_result)        # 跑通接口，得到实际结果

# cases=read_case('test_case_api.xlsx','register')
# for case in cases:           # 直接在上面获取到的表格生成的list数据，循环取出要的单个数据,,得出的数据都是字符串数据类型
#     case_id=case['case_id']
#     url=case['url']           # 定义变量名为url，获取表格里的’url‘的值
#     data=eval(case['data'])       # 获取data的值,,把字符串转成字典
#     expect=eval(case['expect'])    #'''获取expect的值,(他的值不是在表格直接取，而是在上面def read_case下的循环里定义的expect值，也是字符串形式，需要转换成字典)'''
#     # print(url,data,expect)
#     # print(case_id, expect)
#     expect_msg=expect['msg']      # 读到预期结果的msg信息
#     real_result=api_fun(url=url,data=data)
#     real_msg = real_result['msg']  # 获取实际结果的msg信息
#     print('期望结果为:{}'.format(expect_msg))
#     print('实际结果为:{}'.format(real_msg))
#     # print(case_id,expect_msg,real_msg)
#     # if expect_msg==real_msg:
#     #     print('passed!')
#     # else:
#     #     print('fail')
#
#     if expect_msg==real_msg:
#         print('第{}条用例通过!'.format(case_id))
#         final_re='passed'
#     else:
#         print('第{}条用例不通过!'.format(case_id))
#         final_re='failed'
#     write_result('test_case_api.xlsx','register',case_id+1,8,final_re)      #    比对预期结果和实际结果后，写入结果到测试用例
#     print('*'*20)


# 封装成一个函数
# def execute_fun(filename,sheetname,):        # 测试执行
#     cases=read_case(filename,sheetname)
#     for case in cases:           # 直接在上面获取到的表格生成的list数据，循环取出要的单个数据,,得出的数据都是字符串数据类型
#         case_id=case['case_id']
#         url=case['url']           # 定义变量名为url，获取表格里的’url‘的值
#         data=eval(case['data'])       # 获取data的值,,把字符串转成字典
#         expect=eval(case['expect'])    #'''获取expect的值,(他的值不是在表格直接取，而是在上面def read_case下的循环里定义的expect值，也是字符串形式，需要转换成字典)'''
#         # print(url,data,expect)
#         # print(case_id, expect)
#         expect_msg=expect['msg']      # 读到预期结果的msg信息
#         real_result=api_fun(url=url,data=data)
#         real_msg = real_result['msg']  # 获取实际结果的msg信息
#         print('期望结果为:{}'.format(expect_msg))
#         print('实际结果为:{}'.format(real_msg))
#         # print(case_id,expect_msg,real_msg)
#         # if expect_msg==real_msg:
#         #     print('passed!')
#         # else:
#         #     print('fail')
#
#         if expect_msg==real_msg:
#             print('第{}条用例通过!'.format(case_id))
#             final_re='passed'
#         else:
#             print('第{}条用例不通过!'.format(case_id))
#             final_re='failed'
#         write_result(filename,sheetname,case_id+1,8,final_re)      #    比对预期结果和实际结果后，写入结果到测试用例
#         print('*'*20)
# 调用函数
# execute_fun('test_case_api.xlsx', 'register')        # 跑通注册接口，并写回结果
# execute_fun('test_case_api.xlsx', 'login')          # 跑通登陆接口，并写回结果
# test_case_api.xlsx表格被移出同级别的package后，接口就跑不通了，具体操作见run.py
